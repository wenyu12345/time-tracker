from flask import Blueprint, request, jsonify
from datetime import datetime, timezone
from utils.db import get_db
from utils.role_mapping import get_base_role, BASE_ROLES

bp = Blueprint('time_record', __name__)


def auto_assign_role_from_history(db, user_id, current_description=None):
    """根据历史数据自动为用户分配岗位

    策略：
    1. 如果用户已有基础岗位（user_roles 中有 BASE_ROLES 中的岗位），不做任何修改
    2. 如果没有，先从当前记录的 description 中提取岗位关键词
    3. 再查询最近 30 天的 time_records.description 字段提取岗位关键词
    4. 找到有效的基础岗位后，写入 user_roles

    Args:
        db: 数据库连接
        user_id: 用户ID
        current_description: 当前正在添加/更新的记录 description（优先使用）
    """
    if not user_id:
        return

    # 1. 检查用户是否已有基础岗位
    existing_roles = db.execute("""
        SELECT r.name as role_name
        FROM user_roles ur
        JOIN roles r ON ur.role_id = r.id
        WHERE ur.user_id = ?
    """, (user_id,)).fetchall()

    existing_role_names = [row['role_name'] for row in existing_roles]
    # 如果已有基础岗位，不修改
    has_base_role = any(r in BASE_ROLES for r in existing_role_names)
    if has_base_role:
        return

    # 2. 从当前 description 中提取岗位
    candidate_descriptions = []
    if current_description:
        candidate_descriptions.append(current_description)

    # 3. 查询最近 30 天的 time_records，获取 description 字段
    recent_rows = db.execute("""
        SELECT COALESCE(description, '') as desc_text
        FROM time_records
        WHERE user_id = ?
          AND description IS NOT NULL
          AND description != ''
          AND start_time >= datetime('now', '-30 days')
        ORDER BY start_time DESC
        LIMIT 10
    """, (user_id,)).fetchall()

    for row in recent_rows:
        if row['desc_text'] and row['desc_text'] not in candidate_descriptions:
            candidate_descriptions.append(row['desc_text'])

    # 4. 尝试从这些文本中推断岗位
    inferred_role = None
    for text in candidate_descriptions:
        mapped = get_base_role(text)
        if mapped and mapped in BASE_ROLES:
            inferred_role = mapped
            break

    if not inferred_role:
        # 尝试仅检查是否包含关键字
        for text in candidate_descriptions:
            for keyword, base_role in [
                ('配料', '配料'), ('辊压', '辊压'), ('涂布', '涂布'),
                ('分条', '激光切'), ('激光切', '激光切'), ('职能', '职能'),
                ('物料员', '物料'), ('物料', '物料'), ('主管', '主管'),
                ('领班', '领班'), ('发片', '发片'),
            ]:
                if keyword in text:
                    inferred_role = base_role
                    break
            if inferred_role:
                break

    # 5. 写入 user_roles
    if inferred_role:
        role_row = db.execute(
            'SELECT id FROM roles WHERE name = ?',
            (inferred_role,)
        ).fetchone()
        role_id = None
        if role_row:
            role_id = role_row['id']
        else:
            cursor = db.execute(
                'INSERT INTO roles (name) VALUES (?)',
                (inferred_role,)
            )
            role_id = cursor.lastrowid

        # 清除旧的非基础岗位角色，设置新岗位
        db.execute('DELETE FROM user_roles WHERE user_id = ?', (user_id,))
        db.execute(
            'INSERT INTO user_roles (user_id, role_id) VALUES (?, ?)',
            (user_id, role_id)
        )

@bp.route('/', methods=['GET'])
def get_time_records():
    user_id = request.args.get('user_id')
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    
    if not user_id:
        return jsonify({'error': 'User ID is required'}), 400
    
    db = get_db()
    query = 'SELECT * FROM time_records WHERE user_id = ?'
    params = [user_id]
    
    if start_date:
        # 提取日期部分（YYYY-MM-DD），不管格式如何
        if ' ' in start_date:
            date_part = start_date.split(' ')[0]
        else:
            date_part = start_date.split('T')[0] if 'T' in start_date else start_date
        
        # 如果是仅日期查询，使用DATE函数
        if len(date_part) == 10:
            # 如果同时有结束日期，使用范围查询
            if end_date:
                if ' ' in end_date:
                    end_date_part = end_date.split(' ')[0]
                else:
                    end_date_part = end_date.split('T')[0] if 'T' in end_date else end_date
                
                if len(end_date_part) == 10:
                    # 日期范围查询
                    query += ' AND DATE(start_time) >= ? AND DATE(start_time) <= ?'
                    params.append(date_part)
                    params.append(end_date_part)
                    # 清空end_date，避免重复添加
                    end_date = None
                else:
                    # 开始时间查询
                    query += ' AND start_time >= ?'
                    params.append(start_date.replace(' ', 'T') if ' ' in start_date else start_date)
            else:
                # 单个日期查询
                query += ' AND DATE(start_time) = ?'
                params.append(date_part)
        else:
            # 精确时间查询
            query += ' AND start_time >= ?'
            params.append(start_date.replace(' ', 'T') if ' ' in start_date else start_date)
    
    if end_date:
        # 提取日期部分
        if ' ' in end_date:
            date_part = end_date.split(' ')[0]
        else:
            date_part = end_date.split('T')[0] if 'T' in end_date else end_date
        
        if len(date_part) == 10:
            # 日期范围查询（如果前面没有处理过）
            query += ' AND DATE(start_time) <= ?'
            params.append(date_part)
        else:
            # 精确时间查询
            query += ' AND start_time <= ?'
            params.append(end_date.replace(' ', 'T') if ' ' in end_date else end_date)
    
    query += ' ORDER BY start_time DESC'
    records = db.execute(query, params).fetchall()
    
    result = []
    for record in records:
        record_dict = dict(record)
        if record_dict.get('shift_type') == 'day':
            record_dict['shift_type'] = '白班'
        elif record_dict.get('shift_type') == 'night':
            record_dict['shift_type'] = '夜班'
        result.append(record_dict)
    
    return jsonify(result), 200

def _extract_date_part(time_str):
    """从 start_time 中提取 YYYY-MM-DD（兼容 'YYYY-MM-DD HH:MM:SS' 和 ISO 格式）"""
    if not time_str:
        return None
    t = time_str.strip()
    if ' ' in t:
        return t.split(' ')[0]
    if 'T' in t:
        return t.split('T')[0]
    if len(t) >= 10:
        return t[:10]
    return None


@bp.route('/', methods=['POST'])
def create_time_record():
    """创建/覆盖工时记录
    查重规则：同一用户 + 同一日期（DATE(start_time)）只保留一条记录
    - 该日期已有记录 → UPDATE 覆盖
    - 该日期无记录 → INSERT 新建
    """
    data = request.get_json()
    user_id = data.get('user_id')
    project_id = data.get('project_id')
    start_time = data.get('start_time')
    end_time = data.get('end_time')
    duration = data.get('duration')
    shift_type = data.get('shift_type')
    is_leave = data.get('is_leave', False)
    description = data.get('description')
    is_early_off = data.get('is_early_off', False)

    if not user_id or not start_time:
        return jsonify({'error': 'User ID and start time are required'}), 400

    # 统一班别字段
    if shift_type == 'day':
        shift_type = '白班'
    elif shift_type == 'night':
        shift_type = '夜班'

    # 计算时长
    if end_time and not duration:
        try:
            start = datetime.fromisoformat(start_time.replace('Z', '+00:00'))
            end = datetime.fromisoformat(end_time.replace('Z', '+00:00'))
            duration = int((end - start).total_seconds() / 60)  # 转换为分钟
        except Exception as e:
            return jsonify({'error': f'Invalid time format: {str(e)}'}), 400

    # 确定日期（用于查重）
    date_part = _extract_date_part(start_time)
    if not date_part:
        return jsonify({'error': 'Could not determine date from start_time'}), 400

    db = get_db()
    try:
        # 查重：同用户 + 同日期
        existing = db.execute(
            '''SELECT id FROM time_records
               WHERE user_id = ? AND DATE(start_time) = ?
               ORDER BY id DESC LIMIT 1''',
            (user_id, date_part)
        ).fetchone()

        if existing:
            # 有记录：直接 UPDATE 覆盖（覆盖式修改）
            db.execute(
                '''UPDATE time_records
                   SET project_id = ?, start_time = ?, end_time = ?,
                       duration = ?, shift_type = ?, is_leave = ?,
                       description = ?, is_early_off = ?
                   WHERE id = ?''',
                (project_id, start_time, end_time, duration, shift_type,
                 is_leave, description, is_early_off, existing['id'])
            )
            auto_assign_role_from_history(db, user_id, description)
            db.commit()
            return jsonify({
                'message': 'Time record updated (upsert)',
                'id': existing['id'],
                'action': 'update'
            }), 200

        # 无记录：INSERT 新建
        db.execute(
            '''INSERT INTO time_records
               (user_id, project_id, start_time, end_time, duration,
                shift_type, is_leave, description, is_early_off)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)''',
            (user_id, project_id, start_time, end_time, duration,
             shift_type, is_leave, description, is_early_off)
        )
        auto_assign_role_from_history(db, user_id, description)
        db.commit()
        return jsonify({
            'message': 'Time record created successfully',
            'action': 'create'
        }), 201
    except Exception as e:
        db.rollback()
        return jsonify({'error': f'Database error: {str(e)}'}), 500

@bp.route('/start', methods=['POST'])
def start_time_record():
    data = request.get_json()
    user_id = data.get('user_id')
    project_id = data.get('project_id')
    shift_type = data.get('shift_type')
    description = data.get('description')
    
    if not user_id:
        return jsonify({'error': 'User ID is required'}), 400
    
    # 统一班别字段
    if shift_type == 'day':
        shift_type = '白班'
    elif shift_type == 'night':
        shift_type = '夜班'
    
    start_time = datetime.now(timezone.utc).isoformat()
    
    db = get_db()
    db.execute(
        '''INSERT INTO time_records 
        (user_id, project_id, start_time, shift_type, description) 
        VALUES (?, ?, ?, ?, ?)''',
        (user_id, project_id, start_time, shift_type, description)
    )
    db.commit()
    
    return jsonify({'message': 'Time record started', 'start_time': start_time}), 200

@bp.route('/stop', methods=['POST'])
def stop_time_record():
    data = request.get_json()
    record_id = data.get('record_id')
    
    if not record_id:
        return jsonify({'error': 'Record ID is required'}), 400
    
    db = get_db()
    record = db.execute('SELECT * FROM time_records WHERE id = ?', (record_id,)).fetchone()
    
    if not record:
        return jsonify({'error': 'Record not found'}), 404
    
    if record['end_time']:
        return jsonify({'error': 'Record already stopped'}), 400
    
    end_time = datetime.now(timezone.utc).isoformat()
    start = datetime.fromisoformat(record['start_time'].replace('Z', '+00:00'))
    end = datetime.fromisoformat(end_time.replace('Z', '+00:00'))
    duration = int((end - start).total_seconds() / 60)  # 转换为分钟
    
    db.execute(
        'UPDATE time_records SET end_time = ?, duration = ? WHERE id = ?',
        (end_time, duration, record_id)
    )
    db.commit()
    
    return jsonify({'message': 'Time record stopped', 'end_time': end_time, 'duration': duration}), 200

@bp.route('/batch', methods=['POST'])
def batch_update_time_records():
    data = request.get_json()
    user_id = data.get('user_id')
    records = data.get('records')
    
    if not user_id or not records:
        return jsonify({'error': 'User ID and records are required'}), 400
    
    db = get_db()
    
    try:
        batch_descriptions = []
        for record in records:
            record_id = record.get('id')
            if record_id:
                # 更新现有记录
                update_fields = []
                update_params = []

                if 'shift_type' in record:
                    shift_type = record['shift_type']
                    if shift_type == 'day':
                        shift_type = '白班'
                    elif shift_type == 'night':
                        shift_type = '夜班'
                    update_fields.append('shift_type = ?')
                    update_params.append(shift_type)
                if 'is_leave' in record:
                    update_fields.append('is_leave = ?')
                    update_params.append(record['is_leave'])
                if 'start_time' in record:
                    update_fields.append('start_time = ?')
                    update_params.append(record['start_time'])
                if 'end_time' in record:
                    update_fields.append('end_time = ?')
                    update_params.append(record['end_time'])
                    # 重新计算时长
                    if 'start_time' in record:
                        # 解析ISO格式时间，确保时区一致
                        start = datetime.fromisoformat(record['start_time'].replace('Z', '+00:00'))
                        end = datetime.fromisoformat(record['end_time'].replace('Z', '+00:00'))
                        duration = int((end - start).total_seconds() / 60)
                        update_fields.append('duration = ?')
                        update_params.append(duration)

                if update_fields:
                    update_params.append(record_id)
                    update_params.append(user_id)
                    query = f'UPDATE time_records SET {', '.join(update_fields)} WHERE id = ? AND user_id = ?'
                    db.execute(query, update_params)
                # 记录 description 用于岗位推断
                if record.get('description'):
                    batch_descriptions.append(record['description'])
            else:
                # 创建新记录
                start_time = record.get('start_time')
                end_time = record.get('end_time')
                shift_type = record.get('shift_type')
                if shift_type == 'day':
                    shift_type = '白班'
                elif shift_type == 'night':
                    shift_type = '夜班'
                is_leave = record.get('is_leave', False)
                description = record.get('description')
                if description:
                    batch_descriptions.append(description)

                if start_time:
                    duration = None
                    if end_time:
                        # 解析ISO格式时间，确保时区一致
                        start = datetime.fromisoformat(start_time.replace('Z', '+00:00'))
                        end = datetime.fromisoformat(end_time.replace('Z', '+00:00'))
                        duration = int((end - start).total_seconds() / 60)

                    db.execute(
                        '''INSERT INTO time_records 
                        (user_id, start_time, end_time, duration, shift_type, is_leave, description) 
                        VALUES (?, ?, ?, ?, ?, ?, ?)''',
                        (user_id, start_time, end_time, duration, shift_type, is_leave, description)
                    )

        # 批量处理完所有记录后，统一做一次岗位推断
        combined_desc = ' '.join([d for d in batch_descriptions if d]) if batch_descriptions else None
        auto_assign_role_from_history(db, user_id, combined_desc)
        db.commit()
        return jsonify({'message': 'Batch update successful'}), 200
    except Exception as e:
        db.rollback()
        return jsonify({'error': str(e)}), 500

@bp.route('/<int:record_id>', methods=['PUT'])
def update_time_record(record_id):
    """更新单个工时记录"""
    data = request.get_json()
    user_id = request.args.get('user_id')
    
    if not user_id:
        return jsonify({'error': 'User ID is required'}), 400
    
    db = get_db()
    
    # 检查记录是否存在
    record = db.execute('SELECT * FROM time_records WHERE id = ? AND user_id = ?', (record_id, user_id)).fetchone()
    if not record:
        return jsonify({'error': 'Record not found'}), 404
    
    # 准备更新字段
    update_fields = []
    update_params = []
    
    if 'shift_type' in data:
        shift_type = data['shift_type']
        if shift_type == 'day':
            shift_type = '白班'
        elif shift_type == 'night':
            shift_type = '夜班'
        update_fields.append('shift_type = ?')
        update_params.append(shift_type)
    if 'is_leave' in data:
        update_fields.append('is_leave = ?')
        update_params.append(data['is_leave'])
    if 'start_time' in data:
        update_fields.append('start_time = ?')
        update_params.append(data['start_time'])
    if 'end_time' in data:
        update_fields.append('end_time = ?')
        update_params.append(data['end_time'])
    if 'duration' in data:
        update_fields.append('duration = ?')
        update_params.append(data['duration'])
    if 'description' in data:
        update_fields.append('description = ?')
        update_params.append(data['description'])
    
    # 如果没有要更新的字段，直接返回成功
    if not update_fields:
        return jsonify({'message': 'No fields to update'}), 200
    
    # 重新计算时长（如果提供了start_time和end_time）
    if 'start_time' in data and 'end_time' in data:
        start = datetime.fromisoformat(data['start_time'].replace('Z', '+00:00'))
        end = datetime.fromisoformat(data['end_time'].replace('Z', '+00:00'))
        duration = int((end - start).total_seconds() / 60)
        update_fields.append('duration = ?')
        update_params.append(duration)
    
    # 执行更新
    update_params.append(record_id)
    update_params.append(user_id)
    query = f'UPDATE time_records SET {', '.join(update_fields)} WHERE id = ? AND user_id = ?'
    db.execute(query, update_params)

    # 自动根据历史数据分配岗位（仅在用户尚未设置基础岗位时生效）
    new_description = data.get('description', record.get('description'))
    auto_assign_role_from_history(db, user_id, new_description)
    db.commit()

    return jsonify({'message': 'Time record updated successfully'}), 200

@bp.route('/<int:record_id>', methods=['DELETE'])
def delete_time_record(record_id):
    user_id = request.args.get('user_id')
    if not user_id:
        return jsonify({'error': 'User ID is required'}), 400
    
    db = get_db()
    db.execute('DELETE FROM time_records WHERE id = ? AND user_id = ?', (record_id, user_id))
    db.commit()
    
    return jsonify({'message': 'Time record deleted successfully'}), 200

@bp.route('/import-excel/sheets', methods=['POST'])
def get_excel_sheets():
    """获取Excel文件中的所有工作表名称"""
    from flask import request
    from openpyxl import load_workbook
    
    if 'file' not in request.files:
        return jsonify({'error': 'No file uploaded'}), 400
    
    file = request.files['file']
    
    if file.filename == '':
        return jsonify({'error': 'No selected file'}), 400
    
    if not file.filename.endswith('.xlsx'):
        return jsonify({'error': 'Only Excel files (.xlsx) are supported'}), 400
    
    try:
        wb = load_workbook(file)
        sheets = wb.sheetnames
        return jsonify({'sheets': sheets}), 200
    except Exception as e:
        return jsonify({'error': f'Failed to read Excel file: {str(e)}'}), 500

@bp.route('/import-excel', methods=['POST'])
def import_excel_time_records():
    """批量导入Excel排班工时数据，包含查重逻辑"""
    from flask import request
    from openpyxl import load_workbook
    from datetime import datetime, timedelta
    import os
    
    # 检查是否有文件上传
    if 'file' not in request.files:
        return jsonify({'error': 'No file uploaded'}), 400
    
    file = request.files['file']
    
    if file.filename == '':
        return jsonify({'error': 'No selected file'}), 400
    
    # 检查文件类型
    if not file.filename.endswith('.xlsx'):
        return jsonify({'error': 'Only Excel files (.xlsx) are supported'}), 400
    
    # 获取用户ID（从请求参数或表单中获取）
    user_id = request.form.get('user_id')
    if not user_id:
        return jsonify({'error': 'User ID is required'}), 400
    
    try:
        user_id = int(user_id)
    except ValueError:
        return jsonify({'error': 'Invalid user ID'}), 400
    
    # 获取工作表名称（可选，默认为活动工作表）
    sheet_name = request.form.get('sheet_name')
    
    db = get_db()
    
    # 检查用户是否存在
    if not db.execute('SELECT id FROM users WHERE id = ?', (user_id,)).fetchone():
        return jsonify({'error': 'User not found'}), 404
    
    try:
        # 加载Excel文件
        wb = load_workbook(file)
        
        # 根据指定的工作表名称选择工作表
        if sheet_name and sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
        else:
            sheet = wb.active
        
        # 用于统计
        total_rows = 0
        success_count = 0
        duplicate_count = 0
        error_count = 0
        errors = []
        
        # 定义列索引（根据用户提供的Excel格式）
        # 列顺序：日期, 班次, 上班时间, 下班时间, 工时(小时)
        date_col = 0
        shift_type_col = 1
        start_time_col = 2
        end_time_col = 3
        hours_col = 4
        
        # 跳过表头，从第二行开始读取
        for row in sheet.iter_rows(min_row=2, values_only=True):
            total_rows += 1
            
            try:
                # 解析日期
                record_date = row[date_col]
                date_str = None
                
                # 情况1: datetime对象
                if isinstance(record_date, datetime):
                    date_str = record_date.strftime('%Y-%m-%d')
                
                # 情况2: 字符串格式
                elif isinstance(record_date, str):
                    # 尝试解析不同格式的日期字符串
                    date_formats = [
                        '%Y-%m-%d', 
                        '%Y/%m/%d', 
                        '%d-%m-%Y', 
                        '%d/%m/%Y', 
                        '%Y年%m月%d日',
                        '%Y年%m月%d',
                        '%Y.%m.%d',
                        '%Y年%m月',
                        '%Y-%m'
                    ]
                    for fmt in date_formats:
                        try:
                            date_str = datetime.strptime(record_date.strip(), fmt).strftime('%Y-%m-%d')
                            break
                        except ValueError:
                            continue
                    
                    # 如果上述格式都不匹配，尝试解析"X月X日"格式（缺少年份）
                    if not date_str:
                        import re
                        # 匹配"5月1日"或"5月10日"格式
                        match = re.match(r'(\d{1,2})月(\d{1,2})日?$', record_date.strip())
                        if match:
                            month = int(match.group(1))
                            day = int(match.group(2))
                            # 使用当前年份
                            current_year = datetime.now().year
                            try:
                                date_str = f'{current_year:04d}-{month:02d}-{day:02d}'
                            except ValueError:
                                pass
                
                # 情况3: 数字格式（Excel日期序列号）
                elif isinstance(record_date, (int, float)):
                    # Excel日期序列号转换
                    # Excel epoch is 1899-12-30
                    excel_epoch = datetime(1899, 12, 30)
                    try:
                        # 添加天数（处理小数部分表示时间）
                        date_obj = excel_epoch + timedelta(days=int(record_date))
                        date_str = date_obj.strftime('%Y-%m-%d')
                    except OverflowError:
                        pass
                
                # 情况4: openpyxl的datetime对象
                elif hasattr(record_date, 'year') and hasattr(record_date, 'month') and hasattr(record_date, 'day'):
                    try:
                        date_str = f'{record_date.year:04d}-{record_date.month:02d}-{record_date.day:02d}'
                    except:
                        pass
                
                if not date_str:
                    errors.append(f'第{total_rows+1}行：无法解析日期格式 (值: {record_date}, 类型: {type(record_date).__name__})')
                    error_count += 1
                    continue
                
                # 解析班次
                shift_type = row[shift_type_col]
                if not shift_type:
                    shift_type = '白班'  # 默认白班
                shift_type = str(shift_type).strip()
                
                # 处理放假/请假情况
                is_leave = False
                if shift_type in ['放假', '请假', '休息', '调休']:
                    is_leave = True
                    # 保持原班次类型，不统一转换为请假
                    # 放假、请假、休息、调休都标记为请假状态，但保留原始班次类型
                
                # 解析上班时间和下班时间
                start_time_str = row[start_time_col]
                end_time_str = row[end_time_col]
                
                # 处理时间格式
                def parse_time(time_val):
                    if isinstance(time_val, datetime):
                        return time_val.strftime('%H:%M:%S')
                    elif isinstance(time_val, str):
                        # 移除可能的空格和多余字符
                        time_val = time_val.strip()
                        # 尝试解析
                        time_formats = ['%H:%M:%S', '%H:%M', '%I:%M %p']
                        for fmt in time_formats:
                            try:
                                return datetime.strptime(time_val, fmt).strftime('%H:%M:%S')
                            except ValueError:
                                continue
                    return None
                
                parsed_start_time = parse_time(start_time_str)
                parsed_end_time = parse_time(end_time_str)
                
                # 如果没有提供具体时间，根据班次类型使用默认时间
                if not parsed_start_time or not parsed_end_time:
                    if shift_type == '夜班':
                        parsed_start_time = '20:00:00'
                        parsed_end_time = '06:00:00'
                    else:
                        parsed_start_time = '08:00:00'
                        parsed_end_time = '18:00:00'
                
                # 构建完整的开始时间和结束时间
                start_time_full = f'{date_str}T{parsed_start_time}'
                
                # 处理夜班跨天情况
                if shift_type == '夜班' and parsed_end_time < parsed_start_time:
                    # 夜班结束时间是第二天
                    next_day = (datetime.strptime(date_str, '%Y-%m-%d') + timedelta(days=1)).strftime('%Y-%m-%d')
                    end_time_full = f'{next_day}T{parsed_end_time}'
                else:
                    end_time_full = f'{date_str}T{parsed_end_time}'
                
                # 解析工时（小时）
                hours = row[hours_col]
                if hours is None:
                    # 计算工时
                    try:
                        start_dt = datetime.strptime(start_time_full, '%Y-%m-%dT%H:%M:%S')
                        end_dt = datetime.strptime(end_time_full, '%Y-%m-%dT%H:%M:%S')
                        duration = int((end_dt - start_dt).total_seconds() / 60)
                    except:
                        duration = 480  # 默认8小时
                else:
                    try:
                        duration = int(float(hours) * 60)
                    except:
                        duration = 480
                
                # 查重逻辑：检查同一用户在同一日期是否已有工时记录
                existing_record = db.execute(
                    '''SELECT id FROM time_records 
                    WHERE user_id = ? AND DATE(start_time) = ?''',
                    (user_id, date_str)
                ).fetchone()
                
                if existing_record:
                    # 覆盖更新现有记录（以最后一次上传为主）
                    db.execute(
                        '''UPDATE time_records 
                        SET shift_type = ?, start_time = ?, end_time = ?, duration = ?, is_leave = ?
                        WHERE id = ?''',
                        (shift_type, start_time_full, end_time_full, duration, is_leave, existing_record[0])
                    )
                    success_count += 1
                    continue
                
                # 计算时长（分钟）- 如果没有提供工时，自动计算
                if duration is None or duration == 0:
                    try:
                        start_dt = datetime.strptime(start_time_full, '%Y-%m-%dT%H:%M:%S')
                        end_dt = datetime.strptime(end_time_full, '%Y-%m-%dT%H:%M:%S')
                        duration = int((end_dt - start_dt).total_seconds() / 60)
                    except:
                        duration = 480
                
                # 插入工时记录
                db.execute(
                    '''INSERT INTO time_records 
                    (user_id, start_time, end_time, duration, shift_type, is_leave, description) 
                    VALUES (?, ?, ?, ?, ?, ?, ?)''',
                    (user_id, start_time_full, end_time_full, duration, shift_type, is_leave, 'Excel导入')
                )
                
                success_count += 1
                
            except Exception as e:
                error_count += 1
                errors.append(f'第{total_rows+1}行：{str(e)}')

        # Excel 导入完成后，为该用户自动推断一次岗位
        auto_assign_role_from_history(db, user_id, 'Excel导入')
        db.commit()

        return jsonify({
            'message': 'Excel import completed',
            'total_rows': total_rows,
            'success_count': success_count,
            'duplicate_count': duplicate_count,
            'error_count': error_count,
            'errors': errors[:10]  # 最多返回10条错误信息
        }), 200

    except Exception as e:
        db.rollback()
        return jsonify({'error': f'Failed to import Excel: {str(e)}'}), 500


@bp.route('/cleanup-duplicates', methods=['POST'])
def cleanup_duplicate_time_records():
    """清理重复的工时记录，保留每个用户每天的最后一条记录"""
    data = request.get_json()
    user_id = data.get('user_id')
    
    if not user_id:
        return jsonify({'error': 'User ID is required'}), 400
    
    try:
        user_id = int(user_id)
    except ValueError:
        return jsonify({'error': 'Invalid user ID'}), 400
    
    db = get_db()
    
    if not db.execute('SELECT id FROM users WHERE id = ?', (user_id,)).fetchone():
        return jsonify({'error': 'User not found'}), 404
    
    try:
        duplicate_query = '''
            SELECT id FROM time_records 
            WHERE user_id = ? 
            AND DATE(start_time) IN (
                SELECT DATE(start_time) 
                FROM time_records 
                WHERE user_id = ? 
                GROUP BY DATE(start_time) 
                HAVING COUNT(*) > 1
            )
            AND id NOT IN (
                SELECT MAX(id) 
                FROM time_records 
                WHERE user_id = ? 
                GROUP BY DATE(start_time)
            )
        '''
        
        duplicates = db.execute(duplicate_query, (user_id, user_id, user_id)).fetchall()
        
        if not duplicates:
            return jsonify({
                'message': 'No duplicate records found',
                'deleted_count': 0
            }), 200
        
        duplicate_ids = [record['id'] for record in duplicates]
        delete_query = 'DELETE FROM time_records WHERE id IN ({})'.format(
            ','.join('?' * len(duplicate_ids))
        )
        
        cursor = db.execute(delete_query, duplicate_ids)
        db.commit()
        
        return jsonify({
            'message': 'Duplicate records cleaned up successfully',
            'deleted_count': len(duplicate_ids)
        }), 200
        
    except Exception as e:
        db.rollback()
        return jsonify({'error': f'Failed to cleanup duplicates: {str(e)}'}), 500

@bp.route('/find-duplicates', methods=['GET'])
def find_duplicate_time_records():
    """查找重复的工时记录"""
    user_id = request.args.get('user_id')
    
    if not user_id:
        return jsonify({'error': 'User ID is required'}), 400
    
    try:
        user_id = int(user_id)
    except ValueError:
        return jsonify({'error': 'Invalid user ID'}), 400
    
    db = get_db()
    
    if not db.execute('SELECT id FROM users WHERE id = ?', (user_id,)).fetchone():
        return jsonify({'error': 'User not found'}), 404
    
    try:
        duplicate_query = '''
            SELECT 
                id, 
                user_id, 
                start_time, 
                end_time, 
                duration, 
                shift_type,
                DATE(start_time) as date_only,
                COUNT(*) as record_count
            FROM time_records 
            WHERE user_id = ? 
            GROUP BY user_id, DATE(start_time)
            HAVING COUNT(*) > 1
            ORDER BY DATE(start_time) DESC
        '''
        
        duplicates = db.execute(duplicate_query, (user_id,)).fetchall()
        
        result = []
        for record in duplicates:
            date_records = db.execute(
                '''SELECT * FROM time_records
                   WHERE user_id = ? AND DATE(start_time) = ?
                   ORDER BY id DESC''',
                (user_id, record['date_only'])
            ).fetchall()

            processed_records = []
            for r in date_records:
                record_dict = dict(r)
                if record_dict.get('shift_type') == 'day':
                    record_dict['shift_type'] = '白班'
                elif record_dict.get('shift_type') == 'night':
                    record_dict['shift_type'] = '夜班'
                processed_records.append(record_dict)

            result.append({
                'date': record['date_only'],
                'record_count': record['record_count'],
                'records': processed_records
            })

        total_duplicate_records = sum(r['record_count'] for r in result)
        
        return jsonify({
            'duplicate_dates': result,
            'total_duplicate_dates': len(result),
            'total_duplicate_records': total_duplicate_records
        }), 200
        
    except Exception as e:
        return jsonify({'error': f'Failed to find duplicates: {str(e)}'}), 500